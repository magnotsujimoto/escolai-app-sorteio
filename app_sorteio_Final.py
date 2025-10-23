import streamlit as st
import pandas as pd
import random
import os, pathlib
import time
import io
import tempfile
from collections import defaultdict

# ============================================================
# BLOCO 1: CONFIGURAÇÃO INICIAL + SESSION STATE (CORRIGIDO)
# ============================================================
st.set_page_config(
    page_title="Aplicativo de Distribuição de Avaliadores | Programa Escolaí",
    page_icon="🚀",
    layout="centered",
)

# -------------------------------
# Inicialização de chaves do session_state
# -------------------------------
if "page" not in st.session_state:
    st.session_state["page"] = "home"

if "last_result" not in st.session_state:
    st.session_state["last_result"] = None

if "sorteio_count" not in st.session_state:
    st.session_state["sorteio_count"] = 0

if "uploaded_file" not in st.session_state:
    st.session_state["uploaded_file"] = None

if "df_escolas" not in st.session_state:
    st.session_state["df_escolas"] = None

if "df_avaliadores" not in st.session_state:
    st.session_state["df_avaliadores"] = None

if "modo_distribuicao" not in st.session_state:
    st.session_state["modo_distribuicao"] = "diversidade"  # valores: "diversidade" ou "mesmo_estado"

# -------------------------------
# Caminho para logo
# -------------------------------
caminho_logo = os.path.join(os.path.dirname(__file__), "logo-escolai.png")
caminho_logo = pathlib.Path(caminho_logo).resolve().as_posix()

# ============================================================
# BLOCO 2: CSS / DESIGN
# ============================================================
st.markdown(
    """
    <style>
    /* ============================== GERAL ============================== */
    html, body, [class*="css"] {
        font-family: "Segoe UI", Arial, sans-serif;
        font-size: 15px;
        color: #333;
    }
    h1, h2, h3, h4 {
        font-weight: 600;
        margin-top: 1.2em;
        margin-bottom: 0.8em;
        color: #2d0a4e;
    }
    p {
        margin: 4px 0;
    }

    /* ============================== LOGO ============================== */
    .logo-escolai {
        display: block;
        margin-left: auto;
        margin-right: auto;
        margin-bottom: 20px;
    }

    /* ============================== BOTÕES ============================== */
    div.stButton > button,
    div.stDownloadButton > button {
        background: #EA7D21 !important;
        color: #fff !important;
        font-weight: 600 !important;
        font-size: 15px !important;
        padding: 12px 28px !important;
        border: none !important;
        border-radius: 6px !important;
        box-shadow: 2px 2px 6px rgba(0,0,0,0.2) !important;
        cursor: pointer !important;
        display: block;
        margin: 20px auto !important;
        text-align: center !important;
    }
    div.stButton > button:hover,
    div.stDownloadButton > button:hover {
        background: #cf6d1a !important;
    }

    /* ============================== BOTÃO SECUNDÁRIO (LILÁS) ============================== */
    div.stButton > button[kind="secondary"] {
        background: #9b59b6 !important;
        color: #fff !important;
    }
    div.stButton > button[kind="secondary"]:hover {
        background: #8e44ad !important;
    }

    /* ============================== TABELAS ============================== */
    .table-escolai {
        margin-top: 15px;
        margin-bottom: 25px;
        border-collapse: collapse;
        width: 100% !important;
    }
    .table-escolai th, .table-escolai td {
        padding: 8px 10px !important;
        text-align: left !important;
        font-size: 14px !important;
    }
    .table-escolai th {
        background-color: #f3f0fa !important;
        font-weight: 600 !important;
        color: #2d0a4e !important;
    }
    .table-escolai tr:nth-child(even) {
        background-color: #fafafa !important;
    }

    /* Scroll horizontal */
    [data-testid="stDataFrame"] table {
        display: block;
        overflow-x: auto;
        white-space: nowrap;
    }
    [data-testid="stDataFrame"] table td {
        white-space: pre-wrap !important;
        word-wrap: break-word !important;
    }

    /* ============================== CAIXAS ============================== */
    .resumo-box {
        background-color: #f3e8ff;
        padding: 18px;
        border-radius: 8px;
        color: #2d0a4e;
        font-weight: 500;
        line-height: 1.6;
        height: 100%;
    }
    .error-box {
        background-color: #ffe8e8;
        padding: 18px;
        border-radius: 8px;
        color: #c00;
        font-weight: 600;
        line-height: 1.6;
    }
    .warning-box {
        background-color: #fff3cd;
        padding: 18px;
        border-radius: 8px;
        color: #6c5202;
        font-weight: 600;
        line-height: 1.6;
    }

    /* ============================== ABAS ============================== */
    [role="tab"][aria-selected="true"] {
        color: #6a0dad !important;
        font-weight: 600 !important;
    }
    [role="tab"][aria-selected="false"] {
        color: #444 !important;
    }

    /* ============================== ESPAÇAMENTO ============================== */
    .block-container {
        padding-top: 60px;
        padding-bottom: 25px;
        max-width: 1600px;
    }
    .resumo-section { 
        margin-top: 40px;
        margin-bottom: 40px; 
    }
    .table-section { 
        margin-top: 40px; 
    }

    /* ============================== UPLOAD ============================== */
    div[data-testid="stFileUploadDropzone"] section {
        opacity: 0;
        height: 0;
        overflow: hidden;
    }
    div[data-testid="stFileUploadDropzone"]::before {
        content: "Arraste e solte o arquivo aqui ou clique em 'Procurar Arquivo'";
        display: block;
        text-align: center;
        padding: 12px;
        color: #333;
        font-size: 14px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# BLOCO 3: CABEÇALHO
# ============================================================
col1, col2, col3, col4, col5, col6, col7 = st.columns([2, 1, 2, 2, 2, 1, 2])
with col4:
    if os.path.exists(caminho_logo):
        st.image(caminho_logo, width=200)

st.markdown(
    """
    <div style="text-align: center; margin-bottom:40px; margin-top:10px;">
        <h1 style="color:#472977; font-size:24px; font-weight:bold; margin:0; line-height:1.3;">
            Aplicativo de Distribuição de Avaliadores <br> Programa Escolaí
        </h1>
    </div>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# BLOCO 5: FUNÇÕES AUXILIARES
# ============================================================


# 5.1) Normalização de DataFrames do editor
def _normalize_editor_df(df_raw, checkbox_col=None, numeric_cols=None):
    """
    Normaliza DataFrame retornado pelo st.data_editor
    """
    df = df_raw.copy()

    if checkbox_col and checkbox_col in df.columns:
        df[checkbox_col] = df[checkbox_col].astype(bool)

    if numeric_cols:
        for col in numeric_cols:
            if col in df.columns:
                num_col_name = f"{col}_num"
                # Só cria a coluna se ela NÃO existir
                if num_col_name not in df.columns:
                    df[num_col_name] = (
                        pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
                    )
                else:
                    # Se já existe, apenas atualiza os valores
                    df[num_col_name] = (
                        pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
                    )

    return df


# 5.2) Normalização de Estados
STATE_MAP = {
    "São Paulo": "SP",
    "SP": "SP",
    "Rio de Janeiro": "RJ",
    "RJ": "RJ",
    "Espírito Santo": "ES",
    "ES": "ES",
    "Minas Gerais": "MG",
    "MG": "MG",
    "Paraná": "PR",
    "PR": "PR",
    "Ceará": "CE",
    "CE": "CE",
    "Alagoas": "AL",
    "AL": "AL",
    "Acre": "AC",
    "AC": "AC",
    "Mato Grosso do Sul": "MS",
    "MS": "MS",
    "Maranhão": "MA",
    "MA": "MA",
    "Piauí": "PI",
    "PI": "PI",
}


def normalize_state(value):
    """Normaliza nomes de estados para siglas"""
    if pd.isna(value) or str(value).strip() == "":
        return "NÃO_INFORMADO"
    if not isinstance(value, str):
        value = str(value)

    value = value.strip().title()

    return STATE_MAP.get(value, value.upper()[:2] if len(value) >= 2 else value)


# 5.3) Função helper para parâmetros aleatórios
def criar_parametros_aleatorios():
    return pd.DataFrame(
        {
            "Parametro": ["Avaliadores_por_escola", "Escolas_por_avaliador"],
            "Tipo": ["Distribuição", "Distribuição"],
            "Valor": [1, 999],
            "Valor_num": [1, 999],
            "Utilizar este Critério": [True, True],
        }
    )


# ============================================================
# BLOCO 6: LÓGICA DE SORTEIO + PRÉ-CHECAGEM
# ============================================================


# 6.1) Pré-checagem - MODO DIVERSIDADE (Original)
def pre_check_sorteio(
    df_escolas, df_avaliadores, avaliadores_por_escola, escolas_por_avaliador
):
    msgs_criticas, msgs_avisos = [], []

    if df_escolas is None or df_escolas.empty:
        return (True, ["Planilha de Escolas vazia ou inválida"], [], 0)
    if df_avaliadores is None or df_avaliadores.empty:
        return (True, ["Planilha de Avaliadores vazia ou inválida"], [], 0)

    colunas_escolas = ["ID_ESCOLA", "ESCOLA", "ESTADO_ESCOLA"]
    colunas_avaliadores = [
        "ID_AVALIADOR",
        "NOME_AVALIADOR",
        "ESTADO_AVALIADOR",
        "CIDADE_AVALIADOR",
        "EMAIL_AVALIADOR",
    ]

    faltando_e = [c for c in colunas_escolas if c not in df_escolas.columns]
    faltando_a = [c for c in colunas_avaliadores if c not in df_avaliadores.columns]
    if faltando_e:
        return (True, [f"Colunas faltando em Escolas: {', '.join(faltando_e)}"], [], 0)
    if faltando_a:
        return (
            True,
            [f"Colunas faltando em Avaliadores: {', '.join(faltando_a)}"],
            [],
            0,
        )

    total_escolas = len(df_escolas)
    total_avaliadores = len(df_avaliadores)

    if avaliadores_por_escola <= 0:
        return (True, ["Parâmetro Avaliadores_por_escola deve ser > 0"], [], 0)

    capacidade = total_avaliadores * max(1, escolas_por_avaliador)
    demanda = total_escolas * max(1, avaliadores_por_escola)
    if capacidade < demanda:
        msgs_avisos.append(
            f"⚠️ Capacidade possivelmente insuficiente: {capacidade} < {demanda} necessárias"
        )

    escolas_estado = df_escolas["ESTADO_ESCOLA"].value_counts().to_dict()
    avaliadores_estado = df_avaliadores["ESTADO_AVALIADOR"].value_counts().to_dict()
    if not escolas_estado:
        return (True, ["Nenhum estado válido encontrado"], [], 0)

    for estado, qtd_escolas in escolas_estado.items():
        demanda_estado = qtd_escolas * avaliadores_por_escola
        avaliadores_locais = avaliadores_estado.get(estado, 0)
        capacidade_local = avaliadores_locais * escolas_por_avaliador
        avaliadores_externos = sum(
            q for e, q in avaliadores_estado.items() if e != estado
        )
        capacidade_externa_total = avaliadores_externos * escolas_por_avaliador

        if capacidade_local + capacidade_externa_total < demanda_estado:
            msgs_avisos.append(
                f"⚠️ Estado {estado}: demanda {demanda_estado}, capacidade disponível {capacidade_local + capacidade_externa_total}"
            )

    return (True, msgs_criticas, msgs_avisos, capacidade)


# 6.1b) Pré-checagem - MODO MESMO ESTADO (Nova lógica)
def pre_check_sorteio_mesmo_estado(
    df_escolas, df_avaliadores, avaliadores_por_escola, escolas_por_avaliador
):
    """
    Verifica viabilidade priorizando avaliadores do mesmo estado.
    Gera avisos APENAS para inconsistências CRÍTICAS que impedem a distribuição.
    Caso contrário, retorna vazio (sem avisos).
    """
    msgs_criticas, msgs_avisos = [], []

    if df_escolas is None or df_escolas.empty:
        return (True, ["Planilha de Escolas vazia ou inválida"], [], 0)
    if df_avaliadores is None or df_avaliadores.empty:
        return (True, ["Planilha de Avaliadores vazia ou inválida"], [], 0)

    colunas_escolas = ["ID_ESCOLA", "ESCOLA", "ESTADO_ESCOLA"]
    colunas_avaliadores = [
        "ID_AVALIADOR",
        "NOME_AVALIADOR",
        "ESTADO_AVALIADOR",
        "CIDADE_AVALIADOR",
        "EMAIL_AVALIADOR",
    ]

    faltando_e = [c for c in colunas_escolas if c not in df_escolas.columns]
    faltando_a = [c for c in colunas_avaliadores if c not in df_avaliadores.columns]
    if faltando_e:
        return (True, [f"Colunas faltando em Escolas: {', '.join(faltando_e)}"], [], 0)
    if faltando_a:
        return (
            True,
            [f"Colunas faltando em Avaliadores: {', '.join(faltando_a)}"],
            [],
            0,
        )

    total_escolas = len(df_escolas)
    total_avaliadores = len(df_avaliadores)

    if avaliadores_por_escola <= 0:
        return (True, ["Parâmetro Avaliadores_por_escola deve ser > 0"], [], 0)

    # Capacidade total (considerando que pode usar qualquer avaliador)
    capacidade = total_avaliadores * max(1, escolas_por_avaliador)
    demanda = total_escolas * max(1, avaliadores_por_escola)
    
    # ✅ CRÍTICO: Só avisa se capacidade TOTAL for insuficiente
    if capacidade < demanda:
        msgs_avisos.append(
            f"⚠️ Capacidade total insuficiente: {capacidade} slots disponíveis < {demanda} necessários"
        )

    escolas_estado = df_escolas["ESTADO_ESCOLA"].value_counts().to_dict()
    
    if not escolas_estado:
        return (True, ["Nenhum estado válido encontrado"], [], 0)

    # ✅ Verifica apenas se algum estado NÃO pode ser atendido de jeito nenhum
    for estado, qtd_escolas in escolas_estado.items():
        demanda_estado = qtd_escolas * avaliadores_por_escola
        
        # Total de avaliadores disponíveis para este estado
        capacidade_total_para_estado = total_avaliadores * escolas_por_avaliador
        
        # ✅ CRÍTICO: Só avisa se IMPOSSÍVEL atender (nem com todos os avaliadores)
        if capacidade_total_para_estado < demanda_estado:
            msgs_avisos.append(
                f"⚠️ Estado {estado}: impossível atender demanda {demanda_estado} com capacidade disponível {capacidade_total_para_estado}"
            )
    
    # ✅ Não gera avisos informativos sobre % de cobertura local
    # Cards ficam em branco quando não há problemas críticos

    return (True, msgs_criticas, msgs_avisos, capacidade)


# 6.2) Estruturas globais
aloc_escola = {}
aloc_avaliador = {}
escolas_estado_count = defaultdict(lambda: defaultdict(int))
resultados_globais = []


# 6.3) Função utilitária
def _usados_na_escola(id_escola):
    return {
        r["ID_AVALIADOR"] for r in resultados_globais if r["ID_ESCOLA"] == id_escola
    }


# 6.4) Seleção de candidatos - NOVA LÓGICA: Priorizar Mesmo Estado
def candidatos_mesmo_estado(
    esc_row, df_avaliadores_param, escolas_por_avaliador_param
):
    """
    Lógica alternativa: PRIORIZA avaliadores do mesmo estado da escola.
    Quando não há mais disponíveis do mesmo estado, distribui aleatoriamente.
    """
    est_e = esc_row["ESTADO_ESCOLA"]
    id_esc = esc_row["ID_ESCOLA"]
    usados = _usados_na_escola(id_esc)

    cand_mesmo_estado = []
    cand_outros_estados = []

    for _, av in df_avaliadores_param.iterrows():
        ida = av["ID_AVALIADOR"]
        if ida in usados:
            continue

        disponivel = escolas_por_avaliador_param - aloc_avaliador[ida]
        if disponivel <= 0:
            continue

        est_a = av["ESTADO_AVALIADOR"]

        # Separa em duas listas: mesmo estado vs outros estados
        if est_a == est_e:
            cand_mesmo_estado.append((10, "Mesmo_Estado_Prioritario", av))
        else:
            cand_outros_estados.append((1, "Outro_Estado_Aleatorio", av))

    # Retorna primeiro os do mesmo estado, depois os outros
    # Embaralha para dar aleatoriedade dentro de cada grupo
    random.shuffle(cand_mesmo_estado)
    random.shuffle(cand_outros_estados)
    
    return cand_mesmo_estado + cand_outros_estados


# 6.5) Seleção de candidatos - LÓGICA ORIGINAL: Diversidade
def candidatos_por_camada(
    esc_row, camada, df_avaliadores_param, escolas_por_avaliador_param
):
    est_e = esc_row["ESTADO_ESCOLA"]
    id_esc = esc_row["ID_ESCOLA"]
    usados = _usados_na_escola(id_esc)

    avaliadores_atuais = [r for r in resultados_globais if r["ID_ESCOLA"] == id_esc]
    estados_atuais = {r["ESTADO_AVALIADOR"] for r in avaliadores_atuais}

    cand = []
    for _, av in df_avaliadores_param.iterrows():
        ida = av["ID_AVALIADOR"]
        if ida in usados:
            continue

        disponivel = escolas_por_avaliador_param - aloc_avaliador[ida]
        if disponivel <= 0:
            continue

        est_a = av["ESTADO_AVALIADOR"]
        rep_estado = escolas_estado_count[ida][est_e]
        total_escolas_avaliador = aloc_avaliador[ida]

        if len(estados_atuais) == 1 and est_a in estados_atuais and est_a == est_e:
            continue

        if camada == 1:
            if est_a != est_e and rep_estado == 0:
                cand.append((4, "Critério_Diversidade", av))
        elif camada == 2:
            pct_mesmo_estado = (
                (rep_estado / total_escolas_avaliador * 100)
                if total_escolas_avaliador > 0
                else 0
            )
            if est_a == est_e and pct_mesmo_estado < 50:
                cand.append((2, "Critério_MesmoEstado", av))
            elif est_a != est_e:
                cand.append((2, "Critério_SemiDiversidade", av))
        elif camada == 3:
            pct_mesmo_estado = (
                (rep_estado / total_escolas_avaliador * 100)
                if total_escolas_avaliador > 0
                else 0
            )
            if est_a == est_e and pct_mesmo_estado >= 50:
                cand.append((1, "Diversidade_Comprometida", av))
            else:
                cand.append((1, "Limite de diversidade", av))

    cand.sort(key=lambda x: x[0], reverse=True)
    return cand


# 6.6) Função principal
def sortear_avaliadores(df_escolas_in, df_avaliadores_in, dist_params):
    global aloc_escola, aloc_avaliador, escolas_estado_count, resultados_globais

    if df_escolas_in is None or df_escolas_in.empty:
        return (pd.DataFrame(), 0, 0, 0, [], [], ["Planilha de Escolas vazia"], [], 0)
    if df_avaliadores_in is None or df_avaliadores_in.empty:
        return (
            pd.DataFrame(),
            0,
            0,
            0,
            [],
            [],
            ["Planilha de Avaliadores vazia"],
            [],
            0,
        )
    if dist_params is None or dist_params.empty:
        return (pd.DataFrame(), 0, 0, 0, [], [], ["Parâmetros não configurados"], [], 0)

    inicio = time.time()
    df_escolas_local = df_escolas_in.copy().sample(frac=1).reset_index(drop=True)
    df_avaliadores_local = df_avaliadores_in.copy()

    df_escolas_local["ESTADO_ESCOLA"] = df_escolas_local["ESTADO_ESCOLA"].apply(
        normalize_state
    )
    df_avaliadores_local["ESTADO_AVALIADOR"] = df_avaliadores_local[
        "ESTADO_AVALIADOR"
    ].apply(normalize_state)

    dist_dict = {
        row["Parametro"]: int(row.get("Valor_num", 0) or 0)
        for _, row in dist_params.iterrows()
        if row.get("Utilizar este Critério")
    }
    avaliadores_por_escola = dist_dict.get("Avaliadores_por_escola", 0)
    escolas_por_avaliador = dist_dict.get("Escolas_por_avaliador", 0)

    # ✅ Usa a pré-checagem correta conforme o modo selecionado
    modo = st.session_state.get("modo_distribuicao", "diversidade")
    
    if modo == "mesmo_estado":
        _, msgs_criticas, msgs_avisos, capacidade = pre_check_sorteio_mesmo_estado(
            df_escolas_local,
            df_avaliadores_local,
            avaliadores_por_escola,
            escolas_por_avaliador,
        )
    else:
        _, msgs_criticas, msgs_avisos, capacidade = pre_check_sorteio(
            df_escolas_local,
            df_avaliadores_local,
            avaliadores_por_escola,
            escolas_por_avaliador,
        )

    total_esperado = len(df_escolas_local) * max(1, avaliadores_por_escola)
    resultados_globais = []
    aloc_escola.clear()
    aloc_avaliador.clear()
    escolas_estado_count.clear()

    for _, esc in df_escolas_local.iterrows():
        aloc_escola[esc["ID_ESCOLA"]] = 0
    for _, av in df_avaliadores_local.iterrows():
        aloc_avaliador[av["ID_AVALIADOR"]] = 0

    try:
        for _, esc in df_escolas_local.iterrows():
            tentativas = 0
            while (
                aloc_escola[esc["ID_ESCOLA"]] < avaliadores_por_escola
                and tentativas < len(df_avaliadores_local) * 3
            ):
                tentativas += 1
                
                # ✅ NOVA LÓGICA: verifica qual modo usar
                modo = st.session_state.get("modo_distribuicao", "diversidade")
                
                if modo == "mesmo_estado":
                    # Usa a lógica de priorizar mesmo estado
                    candidatos = candidatos_mesmo_estado(
                        esc, df_avaliadores_local, escolas_por_avaliador
                    )
                else:
                    # Usa a lógica original de diversidade (camadas 1, 2, 3)
                    candidatos = []
                    for camada in [1, 2, 3]:
                        candidatos = candidatos_por_camada(
                            esc, camada, df_avaliadores_local, escolas_por_avaliador
                        )
                        if candidatos:
                            break
                
                if not candidatos:
                    break

                rank, status, melhor_av = candidatos[0]
                resultados_globais.append(
                    {
                        "Escola": esc["ESCOLA"],
                        "ID_ESCOLA": esc["ID_ESCOLA"],
                        "ESTADO_ESCOLA": esc["ESTADO_ESCOLA"],
                        "ID_AVALIADOR": melhor_av["ID_AVALIADOR"],
                        "NOME_AVALIADOR": melhor_av["NOME_AVALIADOR"],
                        "ESTADO_AVALIADOR": melhor_av["ESTADO_AVALIADOR"],
                        "EMAIL_AVALIADOR": melhor_av["EMAIL_AVALIADOR"],
                        "Criterio_Utilizado": status,
                    }
                )
                aloc_escola[esc["ID_ESCOLA"]] += 1
                aloc_avaliador[melhor_av["ID_AVALIADOR"]] += 1
                escolas_estado_count[melhor_av["ID_AVALIADOR"]][
                    esc["ESTADO_ESCOLA"]
                ] += 1
    except Exception as e:
        msgs_avisos.append(f"⚠️ Erro durante sorteio: {str(e)}")

    tempo_exec = time.time() - inicio
    total_gerado = len(resultados_globais)
    df_resultado = pd.DataFrame(resultados_globais)
    if "Alerta" not in df_resultado.columns:
        df_resultado["Alerta"] = ""

    return (
        df_resultado,
        tempo_exec,
        total_esperado,
        total_gerado,
        ["Estado"],
        ["Estado"],
        msgs_criticas,
        msgs_avisos,
        capacidade,
    )


# 6.7) Wrapper
def executar_sorteio(
    df_escolas, df_avaliadores, ed_dist, avaliadores_por_escola, escolas_por_avaliador
):
    (
        resultado,
        tempo_exec,
        total_esperado,
        total_real,
        criterios_totais,
        criterios_atendidos,
        inconsistencias_criticas,
        inconsistencias_avisos,
        capacidade,
    ) = sortear_avaliadores(df_escolas, df_avaliadores, ed_dist)

    resultado = aplicar_verificacao(resultado)

    st.session_state["last_result"] = {
        "resultado": resultado,
        "tempo_exec": tempo_exec,
        "total_esperado": total_esperado,
        "total_real": total_real,
        "criterios_totais": criterios_totais,
        "criterios_atendidos": criterios_atendidos,
        "capacidade": capacidade,
        "avaliadores_por_escola": avaliadores_por_escola,
        "escolas_por_avaliador": escolas_por_avaliador,
        "avisos": (
            inconsistencias_avisos if isinstance(inconsistencias_avisos, list) else []
        ),
        "version": st.session_state.get("sorteio_count", 0),
        "timestamp": time.time(),
    }

    st.session_state["page"] = "result"
    st.rerun()


# ============================================================
# BLOCO 7: VALIDAÇÃO E MARCAÇÃO DE ALERTAS
# ============================================================


def verificacao_por_escola(df: pd.DataFrame) -> pd.DataFrame:
    """
    Análise do ponto de vista da ESCOLA.
    - Se todos avaliadores são do mesmo estado da escola → alerta crítico.
    - Se nenhum avaliador é do estado da escola → diversidade garantida.
    - Se mistura avaliadores do estado e de fora → marcar avaliador a avaliador.
    """
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Escola",
                "ID_ESCOLA",
                "ESTADO_ESCOLA",
                "ID_AVALIADOR",
                "NOME_AVALIADOR",
                "ESTADO_AVALIADOR",
                "EMAIL_AVALIADOR",
                "Criterio_Utilizado",
                "Alerta",
            ]
        )

    df = df.copy()
    if "Alerta" not in df.columns:
        df["Alerta"] = ""

    for escola_id in df["ID_ESCOLA"].unique():
        mask = df["ID_ESCOLA"] == escola_id
        estado_escola = df.loc[mask, "ESTADO_ESCOLA"].iloc[0]
        estados_avals = set(df.loc[mask, "ESTADO_AVALIADOR"])

        if estados_avals == {estado_escola}:
            df.loc[mask, "Alerta"] = (
                f"⚠️ Todos avaliadores do mesmo estado ({estado_escola})"
            )

        elif estado_escola not in estados_avals:
            df.loc[mask, "Alerta"] = "✓ Diversidade garantida"

        else:
            for idx in df.loc[mask].index:
                if df.at[idx, "ESTADO_AVALIADOR"] == estado_escola:
                    df.at[idx, "Alerta"] = (
                        f"ℹ️ Parte avaliadores do mesmo estado ({estado_escola})"
                    )
                else:
                    df.at[idx, "Alerta"] = "✓ Diversidade garantida"

    return df


def verificacao_por_avaliador(df: pd.DataFrame) -> pd.DataFrame:
    """
    Análise do ponto de vista do AVALIADOR.
    - Se todas as escolas são do mesmo estado do avaliador → alerta crítico.
    - Se nenhuma escola é do estado dele → diversidade garantida.
    - Se mistura escolas do mesmo estado e de fora → marca avaliador a avaliador.
    """
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "ID_AVALIADOR",
                "NOME_AVALIADOR",
                "ESTADO_AVALIADOR",
                "EMAIL_AVALIADOR",
                "Escola",
                "ID_ESCOLA",
                "ESTADO_ESCOLA",
                "Criterio_Utilizado",
                "Alerta",
            ]
        )

    df = df.copy()
    if "Alerta" not in df.columns:
        df["Alerta"] = ""

    for aval_id in df["ID_AVALIADOR"].unique():
        mask = df["ID_AVALIADOR"] == aval_id
        estado_av = df.loc[mask, "ESTADO_AVALIADOR"].iloc[0]
        estados_escolas = set(df.loc[mask, "ESTADO_ESCOLA"])

        if len(estados_escolas) == 1 and estado_av in estados_escolas:
            df.loc[mask, "Alerta"] = f"⚠️ Todas as escolas do mesmo estado ({estado_av})"

        elif estado_av not in estados_escolas:
            df.loc[mask, "Alerta"] = "✓ Diversidade garantida"

        else:
            for idx in df.loc[mask].index:
                if df.at[idx, "ESTADO_ESCOLA"] == estado_av:
                    df.at[idx, "Alerta"] = (
                        f"ℹ️ Parte escolas do mesmo estado ({estado_av})"
                    )
                else:
                    df.at[idx, "Alerta"] = "✓ Diversidade garantida"

    return df


def aplicar_verificacao(df: pd.DataFrame, modo: str = "escola") -> pd.DataFrame:
    """
    Direciona para a verificação correta:
    - modo="escola" → usa verificacao_por_escola
    - modo="avaliador" → usa verificacao_por_avaliador
    """
    if df is None or df.empty:
        return df
    if modo == "avaliador":
        return verificacao_por_avaliador(df)
    return verificacao_por_escola(df)


# ============================================================
# BLOCO 9 + 11: RESULTADOS COMPLETOS (CARDS, BOTÕES, TABELAS) - CORRIGIDO FINAL
# ============================================================
if st.session_state["page"] == "result":
    data = st.session_state.get("last_result", None)
    if not data or "resultado" not in data:
        st.warning("⚠️ Nenhum resultado disponível ainda. Volte e gere a distribuição.")
        st.session_state["page"] = "home"
        st.rerun()

    resultado = pd.DataFrame(data.get("resultado", []))
    if "Alerta" not in resultado.columns:
        resultado["Alerta"] = ""

    # ------------------------------------------------------------
    # PARÂMETROS VISÍVEIS TAMBÉM NA TELA DE RESULTADOS  (mover para o topo)
    # ------------------------------------------------------------
    if (
        "param_dist_df" in st.session_state
        and not st.session_state["param_dist_df"].empty
    ):
        # ✅ SELEÇÃO DO MODO DE DISTRIBUIÇÃO (TELA DE RESULTADOS)
        st.markdown("### 🎯 Modo de Distribuição")
        
        modo_selecionado_result = st.radio(
            "Escolha como distribuir os avaliadores:",
            options=["diversidade", "mesmo_estado"],
            format_func=lambda x: {
                "diversidade": "🌍 Priorizar Diversidade (lógica original)",
                "mesmo_estado": "🏠 Priorizar Mesmo Estado (nova lógica)"
            }[x],
            index=0 if st.session_state.get("modo_distribuicao", "diversidade") == "diversidade" else 1,
            key="radio_modo_distribuicao_result",
            help=(
                "**Priorizar Diversidade:** Tenta distribuir avaliadores de estados diferentes.\n\n"
                "**Priorizar Mesmo Estado:** Prioriza avaliadores do mesmo estado da escola primeiro. "
                "Quando não há mais disponíveis, distribui aleatoriamente."
            )
        )
        
        st.session_state["modo_distribuicao"] = modo_selecionado_result
        
        st.markdown("---")
        
        st.markdown("### ⚙️ Parâmetros de Distribuição (ajuste se desejar)")

        # visão sem coluna técnica
        df_view = st.session_state["param_dist_df"].copy()
        if "Valor_num" in df_view.columns:
            df_view = df_view.drop(columns=["Valor_num"])

        # editor + retorno capturado
        edited_res = st.data_editor(
            df_view,
            key="param_editor_result",
            num_rows="fixed",
            hide_index=True,
            column_config={
                "Parametro": st.column_config.TextColumn("Parametro", disabled=True),
                "Tipo": st.column_config.TextColumn("Tipo", disabled=True),
                "Valor": st.column_config.NumberColumn("Valor", step=1),
                "Utilizar este Critério": st.column_config.CheckboxColumn(
                    "Utilizar este Critério"
                ),
            },
        )

        # sincroniza c/ session_state e remove coluna técnica
        if isinstance(edited_res, (pd.DataFrame, list, dict)):
            df_tmp = pd.DataFrame(edited_res)
            if "Valor_num" in df_tmp.columns:
                df_tmp = df_tmp.drop(columns=["Valor_num"])
            st.session_state["param_dist_df"].loc[:, df_tmp.columns] = df_tmp[
                df_tmp.columns
            ]

        st.markdown(
            "<p style='font-size:13px; color:#555; text-align:center; margin-top:8px;'>"
            "As alterações feitas acima são salvas automaticamente e podem ser usadas ao clicar em "
            "<strong>🎲 Aleatório</strong> ou <strong>🔄 Gerar/Repetir</strong>."
            "</p>",
            unsafe_allow_html=True,
        )

    # ------------------------------------------------------------
    # CARDS DE RESUMO E AVISOS
    # ------------------------------------------------------------
    st.markdown('<div class="resumo-section">', unsafe_allow_html=True)
    col_a, col_b = st.columns([1, 1])

    with col_a:
        tipo_sorteio = (
            "Aleatório"
            if (
                data.get("avaliadores_por_escola", 0) == 1
                and data.get("escolas_por_avaliador", 0) >= 999
            )
            else "Parametrizado"
        )
        st.markdown(
            f"""<div class="resumo-box" style="min-height:200px; height:auto">
            <h4 style="margin-top:0;">📌 Resumo Geral</h4>
            <div style="line-height:2;">
            <strong>Tipo Sorteio:</strong> {tipo_sorteio}<br>
            ⏱️ Tempo: {data.get('tempo_exec', 0):.2f}s<br>
            🎯 Esperado: {data.get('total_esperado', 0)}<br>
            📌 Gerado: {data.get('total_real', 0)}<br>
            📊 Capacidade: {int(data.get('capacidade', 0))}
            </div>
            </div>""",
            unsafe_allow_html=True,
        )

    with col_b:
        avisos = data.get("avisos", [])
        modo_atual = st.session_state.get("modo_distribuicao", "diversidade")
        
        # ✅ MODO DIVERSIDADE: Mostra avisos e observações
        if modo_atual == "diversidade":
            if avisos:
                avisos_html = "<br>".join([f"• {av}" for av in avisos])
                st.markdown(
                    f"""<div class="warning-box" style="margin-bottom:12px; height:auto;">
                    <h4 style="margin-top:0; font-size:14px;">⚠️ Avisos</h4>
                    <div style="font-size:13px; line-height:1.5;">{avisos_html}</div>
                    </div>""",
                    unsafe_allow_html=True,
                )

            df_escola_check = verificacao_por_escola(resultado.copy())
            criticos_total = (df_escola_check["Alerta"].str.contains("⚠️", na=False)).sum()

            if criticos_total > 0:
                st.markdown(
                    f"""<div class="error-box" style="height:auto;">
                    <h4 style="margin-top:0; font-size:14px;">❌ Observações</h4>
                    <div style="font-size:13px;">{criticos_total} casos com baixa diversidade</div>
                    </div>""",
                    unsafe_allow_html=True,
                )
            elif not avisos:
                st.markdown(
                    """<div class="resumo-box" style="height:200px; text-align:center; display:flex; align-items:center; justify-content:center;">
                    <div style="font-size:16px;">✅ Nenhuma observação detectada</div>
                    </div>""",
                    unsafe_allow_html=True,
                )
        
        # ✅ MODO MESMO ESTADO: Cards em branco (apenas erros críticos)
        elif modo_atual == "mesmo_estado":
            # Só mostra avisos se houver erros CRÍTICOS (não informativos)
            if avisos:
                # Filtra apenas avisos críticos (capacidade insuficiente, erros de dados)
                avisos_criticos = [av for av in avisos if "⚠️ Capacidade total insuficiente" in av or "❌" in av]
                
                if avisos_criticos:
                    avisos_html = "<br>".join([f"• {av}" for av in avisos_criticos])
                    st.markdown(
                        f"""<div class="warning-box" style="margin-bottom:12px; height:auto;">
                        <h4 style="margin-top:0; font-size:14px;">⚠️ Avisos</h4>
                        <div style="font-size:13px; line-height:1.5;">{avisos_html}</div>
                        </div>""",
                        unsafe_allow_html=True,
                    )
            # Caso contrário: card fica em branco (sem exibir nada)
    st.markdown("</div>", unsafe_allow_html=True)

    # ------------------------------------------------------------
    # BOTÕES DE AÇÃO
    # ------------------------------------------------------------
    col1, col2, col3, col4, col5, col6, col7 = st.columns([1, 0.7, 1, 1, 1, 0.7, 1])

    with col2:
        if st.button("🗑 Limpar", key="btn_limpar_result"):
            if "last_result" in st.session_state:
                del st.session_state["last_result"]
            st.session_state["page"] = "home"
            st.rerun()

    with col3:
        if st.button("🎲 Aleatório", key="btn_aleatorio_result", type="secondary"):
            df_escolas = st.session_state.get("df_escolas")
            df_avaliadores = st.session_state.get("df_avaliadores")
            if df_escolas is not None and df_avaliadores is not None:
                executar_sorteio(
                    df_escolas, df_avaliadores, criar_parametros_aleatorios(), 1, 999
                )
            else:
                st.error("⚠️ Arquivo Excel inválido ou incompleto.")

    with col4:
        if st.button("🔄 Gerar/Repetir", key="btn_sorteio_result"):
            # normaliza antes de ler Valor_num
            params_raw = st.session_state.get("param_dist_df", pd.DataFrame())
            params = _normalize_editor_df(
                params_raw.copy(),
                checkbox_col="Utilizar este Critério",
                numeric_cols=["Valor"],
            )

            if not params.empty:
                # ✅ CORREÇÃO: Pega apenas parâmetros MARCADOS
                dct = {
                    r["Parametro"]: r.get("Valor_num", None)
                    for _, r in params.iterrows()
                    if r.get("Utilizar este Critério")  # Só pega os marcados
                }
                
                # Verifica se pelo menos 1 parâmetro está marcado e válido
                tem_algum_valido = any(
                    r.get("Utilizar este Critério") and r.get("Valor_num", 0) > 0
                    for _, r in params.iterrows()
                )
                
                if not tem_algum_valido:
                    st.error(
                        "⚠️ Marque e preencha pelo menos um dos parâmetros."
                    )
                else:
                    # Define valores padrão para parâmetros não marcados
                    raw_ape = dct.get("Avaliadores_por_escola", None)
                    raw_epa = dct.get("Escolas_por_avaliador", None)
                    
                    # Se não marcou Avaliadores_por_escola, usa 1 (padrão)
                    if raw_ape is None or pd.isna(raw_ape):
                        avaliadores_por_escola = 1
                    else:
                        avaliadores_por_escola = int(raw_ape)
                    
                    # Se não marcou Escolas_por_avaliador, usa 999 (sem limite)
                    if raw_epa is None or pd.isna(raw_epa):
                        escolas_por_avaliador = 999
                    else:
                        escolas_por_avaliador = int(raw_epa)
                    
                    executar_sorteio(
                        st.session_state.get("df_escolas"),
                        st.session_state.get("df_avaliadores"),
                        params,  # usa params normalizados
                        avaliadores_por_escola,
                        escolas_por_avaliador,
                    )
            else:
                st.error("⚠️ Configure os parâmetros antes de gerar o sorteio.")

        # ------------------------------------------------------------
        # BOTÃO DE DOWNLOAD (reposicionado corretamente)
        # ------------------------------------------------------------
        with col5:
            if not resultado.empty:
                df_escola = aplicar_verificacao(resultado.copy(), "escola")
                df_avaliador = aplicar_verificacao(resultado.copy(), "avaliador")

                cols_aval = [
                    "ID_AVALIADOR",
                    "NOME_AVALIADOR",
                    "ESTADO_AVALIADOR",
                    "EMAIL_AVALIADOR",
                    "Escola",
                    "ID_ESCOLA",
                    "ESTADO_ESCOLA",
                    "Criterio_Utilizado",
                    "Alerta",
                ]
                cols_present = [c for c in cols_aval if c in df_avaliador.columns]
                df_avaliador = df_avaliador[cols_present].sort_values(
                    by=[c for c in ["NOME_AVALIADOR", "Escola"] if c in cols_present]
                )

                # ------------------------------------------------------------
                # CRIAÇÃO DO ARQUIVO EXCEL FORMATADO COM LOG E TERCEIRA ABA
                # ------------------------------------------------------------
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    # 1️⃣ Aba: Resultado por Escola
                    df_escola.to_excel(writer, sheet_name="Por Escola", index=False)

                    # 2️⃣ Aba: Resultado por Avaliador
                    df_avaliador.to_excel(
                        writer, sheet_name="Por Avaliador", index=False
                    )

                    # 3️⃣ Aba: Resumo de Parâmetros e Log
                    df_params = st.session_state.get(
                        "param_dist_df", pd.DataFrame()
                    ).copy()
                    from datetime import datetime

                    datahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                    resumo = {
                        "Campo": [
                            "Tipo de Sorteio",
                            "Tempo de Execução (s)",
                            "Total Esperado",
                            "Total Gerado",
                            "Capacidade",
                            "Data/Hora da Geração",
                        ],
                        "Valor": [
                            (
                                "Aleatório"
                                if (
                                    data.get("avaliadores_por_escola", 0) == 1
                                    and data.get("escolas_por_avaliador", 0) >= 999
                                )
                                else "Parametrizado"
                            ),
                            f"{data.get('tempo_exec', 0):.2f}",
                            data.get("total_esperado", 0),
                            data.get("total_real", 0),
                            int(data.get("capacidade", 0)),
                            datahora,
                        ],
                    }
                    df_resumo = pd.DataFrame(resumo)

                    start_row = 2  # espaço para o título
                    df_resumo.to_excel(
                        writer,
                        sheet_name="Resumo_Parametros",
                        index=False,
                        startrow=start_row,
                    )

                    if not df_params.empty:
                        start_row = len(df_resumo) + 5
                        df_params.to_excel(
                            writer,
                            sheet_name="Resumo_Parametros",
                            index=False,
                            startrow=start_row,
                        )

                # ------------------------------------------------------------
                # APLICA FORMATAÇÃO VISUAL (openpyxl)
                # ------------------------------------------------------------
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                buffer.seek(0)
                wb = load_workbook(buffer)

                def format_sheet(ws):
                    """Aplica cabeçalho formatado, autoajuste e filtro."""
                    header_fill = PatternFill(
                        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                    )
                    thin_border = Border(
                        left=Side(style="thin", color="AAAAAA"),
                        right=Side(style="thin", color="AAAAAA"),
                        top=Side(style="thin", color="AAAAAA"),
                        bottom=Side(style="thin", color="AAAAAA"),
                    )

                    # Cabeçalho
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="000000")
                        cell.fill = header_fill
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        cell.border = thin_border

                    # Autoajuste de largura + bordas
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                            cell.border = thin_border
                        ws.column_dimensions[col_letter].width = max_len + 2

                    ws.auto_filter.ref = ws.dimensions

                # aplica formatação em todas as abas
                for name in wb.sheetnames:
                    ws = wb[name]
                    format_sheet(ws)

                # ------------------------------------------------------------
                # FORMATAÇÃO EXCLUSIVA PARA ABA DE RESUMO
                # ------------------------------------------------------------
                ws_resumo = wb["Resumo_Parametros"]
                lilas = "D9C2E9"  # lilás institucional Escolaí

                # Inserir título grande no topo
                ws_resumo.insert_rows(1)
                ws_resumo["A1"] = (
                    "Relatório de Distribuição de Avaliadores – Programa Escolaí"
                )
                ws_resumo.merge_cells(
                    start_row=1, start_column=1, end_row=1, end_column=4
                )
                ws_resumo["A1"].font = Font(
                    size=15, bold=True, color="4B0082"
                )  # lilás escuro
                ws_resumo["A1"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws_resumo["A1"].fill = PatternFill(
                    start_color=lilas, end_color=lilas, fill_type="solid"
                )
                ws_resumo.row_dimensions[1].height = 28

                # Adiciona borda e ajuste visual
                for cell in ws_resumo[2]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="EAEAEA", end_color="EAEAEA", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")

                # Autoajuste novamente após título
                for col in ws_resumo.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws_resumo.column_dimensions[col_letter].width = max_len + 2

                # ------------------------------------------------------------
                # SALVA E DISPONIBILIZA PARA DOWNLOAD
                # ------------------------------------------------------------
                buffer2 = io.BytesIO()
                wb.save(buffer2)
                buffer2.seek(0)

                from datetime import datetime

                st.download_button(
                    label="📄 Baixar Excel",
                    data=buffer2.getvalue(),
                    file_name=f"resultado_sorteio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_baixar_excel",
                )

    # ------------------------------------------------------------
    # TABELAS DE RESULTADOS
    # ------------------------------------------------------------
    st.markdown(
        '<hr style="border: 0; border-top: 1px solid #ddd; margin: 50px 0 30px 0;">',
        unsafe_allow_html=True,
    )
    st.markdown('<div class="table-section">', unsafe_allow_html=True)
    st.subheader("📊 Resultados Detalhados")
    tab1, tab2 = st.tabs(["Resultado por Escola", "Resultado por Avaliador"])

    # ------------------------------------------------------------
    # CONTEÚDO DAS TABELAS DE RESULTADOS
    # ------------------------------------------------------------
    with tab1:
        if not resultado.empty:
            df_escola = aplicar_verificacao(resultado.copy(), "escola")
            st.dataframe(
                df_escola.sort_values(by=["Escola"]),
                use_container_width=True,
                height=600,
            )
        else:
            st.info("Nenhum resultado disponível.")

    with tab2:
        if not resultado.empty:
            df_avaliador = aplicar_verificacao(resultado.copy(), "avaliador")

            # ordem de colunas focada no avaliador
            cols_aval = [
                "ID_AVALIADOR",
                "NOME_AVALIADOR",
                "ESTADO_AVALIADOR",
                "EMAIL_AVALIADOR",
                "Escola",
                "ID_ESCOLA",
                "ESTADO_ESCOLA",
                "Criterio_Utilizado",
                "Alerta",
            ]
            cols_present = [c for c in cols_aval if c in df_avaliador.columns]
            df_avaliador = df_avaliador[cols_present].sort_values(
                by=[c for c in ["NOME_AVALIADOR", "Escola"] if c in cols_present]
            )

            st.dataframe(
                df_avaliador,
                use_container_width=True,
                height=600,
            )
        else:
            st.info("Nenhum resultado disponível.")

    st.markdown("</div>", unsafe_allow_html=True)

# ============================================================
# BLOCO 10: PÁGINA INICIAL (HOME)
# ============================================================

elif st.session_state["page"] == "home":

    # ------------------------------------------------------------
    # SUB-BLOCO 10.1: UPLOAD DE ARQUIVO
    # ------------------------------------------------------------
    col = st.columns([1, 2, 1])[1]
    with col:
        uploaded_file = st.file_uploader(
            label="",
            type=["xlsx"],
            accept_multiple_files=False,
            label_visibility="collapsed",
            key="uploader_excel_home",
        )
        if uploaded_file is not None:
            st.session_state["uploaded_file"] = uploaded_file

    # Aviso se nenhum arquivo foi carregado
    if st.session_state.get("uploaded_file") is None:
        st.markdown(
            """
            <div style="text-align:center; margin:20px 0; color:#555; font-size:15px;">
                ℹ️ Carregue o arquivo Excel na seção inicial para começar.
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ------------------------------------------------------------
    # SUB-BLOCO 10.2: PROCESSAMENTO DO ARQUIVO E ABAS
    # ------------------------------------------------------------
    if st.session_state.get("uploaded_file") is not None:
        try:
            uploaded_file = st.session_state["uploaded_file"]

            # Carrega arquivo em memória
            bytes_data = uploaded_file.getvalue()
            excel_buffer = io.BytesIO(bytes_data)
            xls = pd.ExcelFile(excel_buffer)

            sheet_names = set(xls.sheet_names)

            # --------------------------------------------------------
            # SUB-BLOCO 10.2.1 – EDITOR DE PARÂMETROS (ANTI-PISCAR)
            # --------------------------------------------------------
            
            # ✅ SELEÇÃO DO MODO DE DISTRIBUIÇÃO
            st.markdown("### 🎯 Modo de Distribuição")
            
            modo_selecionado = st.radio(
                "Escolha como distribuir os avaliadores:",
                options=["diversidade", "mesmo_estado"],
                format_func=lambda x: {
                    "diversidade": "🌍 Priorizar Diversidade (lógica original)",
                    "mesmo_estado": "🏠 Priorizar Mesmo Estado (nova lógica)"
                }[x],
                index=0 if st.session_state.get("modo_distribuicao", "diversidade") == "diversidade" else 1,
                key="radio_modo_distribuicao",
                help=(
                    "**Priorizar Diversidade:** Tenta distribuir avaliadores de estados diferentes.\n\n"
                    "**Priorizar Mesmo Estado:** Prioriza avaliadores do mesmo estado da escola primeiro. "
                    "Quando não há mais disponíveis, distribui aleatoriamente."
                )
            )
            
            # Atualiza o session_state
            st.session_state["modo_distribuicao"] = modo_selecionado
            
            st.markdown("---")  # Linha separadora
            
            st.markdown("### ⚙️ Parâmetros de Distribuição")

            if (
                "param_dist_df" not in st.session_state
                or st.session_state["param_dist_df"].empty
            ):
                # cria tabela vazia para edição manual
                st.session_state["param_dist_df"] = pd.DataFrame(
                    {
                        "Parametro": [
                            "Avaliadores_por_escola",
                            "Escolas_por_avaliador",
                        ],
                        "Tipo": ["Distribuição", "Distribuição"],
                        "Valor": ["", ""],
                        "Utilizar este Critério": [False, False],
                    }
                )

            # 2. Pega referência fixa do DataFrame
            param_ref = st.session_state["param_dist_df"]

            # 3. Exibe o editor apenas uma vez, sem recriar a cada rerun
            if "param_editor_initialized" not in st.session_state:
                st.session_state["param_editor_initialized"] = True
                st.session_state["param_editor_value"] = param_ref.copy()

            edited_df = st.data_editor(
                st.session_state["param_editor_value"],
                key="param_editor_static",
                num_rows="fixed",
                hide_index=True,
                column_config={
                    "Parametro": st.column_config.TextColumn(
                        "Parametro", disabled=True
                    ),
                    "Tipo": st.column_config.TextColumn("Tipo", disabled=True),
                    "Valor": st.column_config.NumberColumn(
                        "Valor",
                        help="Informe um número inteiro (ex.: 1, 2, 3 ...)",
                        step=1,
                    ),
                    "Utilizar este Critério": st.column_config.CheckboxColumn(
                        "Utilizar este Critério"
                    ),
                },
            )

            # 4. Atualiza o DataFrame persistido sem recriar o objeto
            if isinstance(edited_df, (pd.DataFrame, list, dict)):
                try:
                    df_tmp = pd.DataFrame(edited_df)
                    # remove coluna técnica Valor_num, que não deve aparecer no editor
                    if "Valor_num" in df_tmp.columns:
                        df_tmp = df_tmp.drop(columns=["Valor_num"])
                    for col in df_tmp.columns:
                        if col in param_ref.columns:
                            param_ref[col] = df_tmp[col]
                        else:
                            param_ref[col] = df_tmp[col]
                except Exception as e:
                    st.warning(f"⚠️ Erro ao sincronizar parâmetros: {e}")

            # 5. Mensagem informativa
            st.markdown(
                """
                <p style='font-size:13px; color:#555; text-align:center; margin-top:8px;'>
                As alterações feitas acima persistem automaticamente e serão usadas
                ao clicar em <strong>🎲 Aleatório</strong> ou <strong>🔄 Gerar/Repetir</strong>.
                </p>
                """,
                unsafe_allow_html=True,
            )

            # --------------------------------------------------------
            # SUB-BLOCO 10.2.2: VISUALIZAÇÃO DAS ABAS DE PLANILHAS
            # --------------------------------------------------------
            abas = st.tabs(["Avaliadores", "Escolas"])

            # Aba Avaliadores
            if "Avaliadores" in sheet_names:
                df_avaliadores = pd.read_excel(
                    excel_buffer, sheet_name="Avaliadores"
                ).fillna("")
                st.session_state["df_avaliadores"] = df_avaliadores
                with abas[0]:
                    st.subheader("🧑 Avaliadores")
                    st.dataframe(df_avaliadores, use_container_width=True)

            # Aba Escolas
            if "Escolas" in sheet_names:
                excel_buffer.seek(0)
                df_escolas = pd.read_excel(excel_buffer, sheet_name="Escolas").fillna(
                    ""
                )
                st.session_state["df_escolas"] = df_escolas
                with abas[1]:
                    st.subheader("📘 Escolas")
                    st.dataframe(df_escolas, use_container_width=True)

        except Exception as e:
            st.error(f"❌ Erro ao processar arquivo: {e}")

        # ------------------------------------------------------------
        # SUB-BLOCO 10.3: BOTÕES DE AÇÃO (VERSÃO FINAL ESTÁVEL)
        # ------------------------------------------------------------
        c1, c2, c3, c4, c5 = st.columns([0.5, 1, 1, 1, 0.5], gap="small")

        # ----------------------------
        # BOTÃO LIMPAR
        # ----------------------------
        with c2:
            if st.button("🗑 Limpar", key="btn_limpar_home"):
                # Limpa apenas o conteúdo, sem recriar o DataFrame
                if "param_dist_df" in st.session_state:
                    df = st.session_state["param_dist_df"]
                    if "Valor" in df.columns:
                        df["Valor"] = ""
                    if "Utilizar este Critério" in df.columns:
                        df["Utilizar este Critério"] = False

                # Remove resultados anteriores
                if "last_result" in st.session_state:
                    del st.session_state["last_result"]

                st.rerun()

        # --------------------------------------------------------
        # BOTÃO ALEATÓRIO
        # --------------------------------------------------------
        with c3:
            if st.button("🎲 Aleatório", key="btn_aleatorio_home"):
                df_escolas = st.session_state.get("df_escolas")
                df_avaliadores = st.session_state.get("df_avaliadores")

                if df_escolas is not None and df_avaliadores is not None:
                    # ✅ Agora lê o DataFrame persistido (não o editor)
                    params_raw = st.session_state.get("param_dist_df", pd.DataFrame())

                    params = _normalize_editor_df(
                        params_raw.copy(),
                        checkbox_col="Utilizar este Critério",
                        numeric_cols=["Valor"],
                    )

                    tem_parametros_validos = any(
                        row.get("Utilizar este Critério")
                        and row.get("Valor_num", 0) > 0
                        for _, row in params.iterrows()
                    )

                    if tem_parametros_validos:
                        dct = {
                            r["Parametro"]: r.get("Valor_num", None)
                            for _, r in params.iterrows()
                            if r.get("Utilizar este Critério")
                        }
                        ape = int(dct.get("Avaliadores_por_escola", 1))
                        epa = int(dct.get("Escolas_por_avaliador", 999))
                        executar_sorteio(
                            df_escolas, df_avaliadores, params.copy(), ape, epa
                        )
                    else:
                        executar_sorteio(
                            df_escolas,
                            df_avaliadores,
                            criar_parametros_aleatorios(),
                            1,
                            999,
                        )
                else:
                    st.error("⚠️ Arquivo Excel inválido ou incompleto.")

        # --------------------------------------------------------
        # BOTÃO GERAR/REPETIR
        # --------------------------------------------------------
        with c4:
            if st.button("🔄 Gerar/Repetir", key="btn_repetir_home"):
                # ✅ Usa o DataFrame persistido
                params_raw = st.session_state.get("param_dist_df", pd.DataFrame())

                params = _normalize_editor_df(
                    params_raw.copy(),
                    checkbox_col="Utilizar este Critério",
                    numeric_cols=["Valor"],
                )

                if not params.empty:
                    # ✅ CORREÇÃO: Pega apenas parâmetros MARCADOS
                    dct = {
                        r["Parametro"]: r.get("Valor_num", None)
                        for _, r in params.iterrows()
                        if r.get("Utilizar este Critério")  # Só pega os marcados
                    }
                    
                    # Verifica se pelo menos 1 parâmetro está marcado e válido
                    tem_algum_valido = any(
                        r.get("Utilizar este Critério") and r.get("Valor_num", 0) > 0
                        for _, r in params.iterrows()
                    )
                    
                    if not tem_algum_valido:
                        st.error(
                            "⚠️ Marque e preencha pelo menos um dos parâmetros."
                        )
                    else:
                        # Define valores padrão para parâmetros não marcados
                        raw_ape = dct.get("Avaliadores_por_escola", None)
                        raw_epa = dct.get("Escolas_por_avaliador", None)
                        
                        # Se não marcou Avaliadores_por_escola, usa 1 (padrão)
                        if raw_ape is None or pd.isna(raw_ape):
                            avaliadores_por_escola = 1
                        else:
                            avaliadores_por_escola = int(raw_ape)
                        
                        # Se não marcou Escolas_por_avaliador, usa 999 (sem limite)
                        if raw_epa is None or pd.isna(raw_epa):
                            escolas_por_avaliador = 999
                        else:
                            escolas_por_avaliador = int(raw_epa)
                        
                        executar_sorteio(
                            st.session_state.get("df_escolas"),
                            st.session_state.get("df_avaliadores"),
                            params,
                            avaliadores_por_escola,
                            escolas_por_avaliador,
                        )
                else:
                    st.error("⚠️ Configure os parâmetros antes de gerar o sorteio.")
